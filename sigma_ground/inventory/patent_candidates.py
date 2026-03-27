#!/usr/bin/env python3
"""Generate patent candidate spreadsheet for MatterShaper/SSBM IP."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Sheet 1: Patent Candidates Overview ──────────────────────────
ws = wb.active
ws.title = "Patent Candidates"

# Colors
HEADER_FILL = PatternFill('solid', fgColor='1B2A4A')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
SUBHEADER_FILL = PatternFill('solid', fgColor='E8EDF3')
SUBHEADER_FONT = Font(name='Arial', bold=True, size=10)
BODY_FONT = Font(name='Arial', size=10)
BLUE_FONT = Font(name='Arial', size=10, color='0000FF')
PRIORITY_HIGH = PatternFill('solid', fgColor='C6EFCE')
PRIORITY_MED = PatternFill('solid', fgColor='FFEB9C')
PRIORITY_LOW = PatternFill('solid', fgColor='FFC7CE')
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='top', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CURRENCY = '#,##0'

# Column widths
widths = {'A': 5, 'B': 28, 'C': 14, 'D': 14, 'E': 45, 'F': 40, 'G': 35, 'H': 30}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Title
ws.merge_cells('A1:H1')
ws['A1'] = 'MatterShaper / SSBM — Patent Candidate Assessment'
ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1B2A4A')
ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:H2')
ws['A2'] = 'Prepared for: Aaron Rhodes — Confidential — Not Legal Advice'
ws['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')
ws.row_dimensions[2].height = 20

# Headers
headers = ['#', 'Invention', 'Priority', 'Est. Value', 'Core Claim', 'Novel Element', 'Nearest Prior Art', 'Filing Notes']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER
ws.row_dimensions[4].height = 25

# Data rows
candidates = [
    {
        'num': 1,
        'name': 'Push Rendering via Analytic Surface Node Projection',
        'priority': 'HIGH',
        'value': 'High',
        'claim': 'A method for rendering 3D scenes wherein surface elements of analytic geometric primitives compute their own illumination response and project themselves onto a pixel grid without ray-surface intersection tests, wherein a simulated light source acts as a measurement trigger and surface nodes produce no output absent photon activation.',
        'novel': 'No ray tracing. Matter is the subject. Light triggers collapse. "No photon, no render" behavior is architecturally enforced. Physics-aware nodes (not statistical fits).',
        'prior_art': 'Gaussian Splatting (Kerbl et al. 2023) — uses statistical point clouds, no physics. Photon mapping (Jensen 1996) — still ray-based. Point-based rendering (Levoy 2000) — no material identity.',
        'notes': 'FILE FIRST. Combines with #2 for system patent. Provisional ~$320.',
    },
    {
        'num': 2,
        'name': 'Physics-Grounded Unified Material Identity System',
        'priority': 'HIGH',
        'value': 'High',
        'claim': 'A system for representing material properties that stores both optical rendering parameters and atomic physics identity (Z, A, crystal structure, cohesive energy, surface energy) in a unified record, such that the same material drives both visual rendering and physical simulation, with properties decomposed into field-invariant and field-scalable components.',
        'novel': 'Single source of truth for rendering AND simulation. EM/QCD decomposition. Surface energy from broken-bond model feeds both specular rendering and mechanical simulation.',
        'prior_art': 'PBR materials (Disney BRDF) — rendering only. FEM material databases — simulation only. No unified system exists that derives both from atomic identity.',
        'notes': 'Best filed as system patent WITH #1. Together: "Physics-aware rendering system."',
    },
    {
        'num': 3,
        'name': 'Sigma Signature Object Representation',
        'priority': 'MEDIUM',
        'value': 'Medium',
        'claim': 'A method for representing 3D objects as ordered collections of analytic quadric primitives with physics-aware material identifiers, enabling lossless geometric scaling, rapid instantiation, and physically accurate simulation without mesh tessellation or polygon approximation.',
        'novel': 'Objects stored as math equations (quadrics), not meshes. Lossless at any scale. Material references carry atomic identity. Layer-based with semantic labels.',
        'prior_art': 'CSG (Constructive Solid Geometry) — uses boolean ops, no material physics. glTF/USD formats — mesh-based, no analytic primitives. STEP/IGES — CAD, no rendering integration.',
        'notes': 'Weaker standalone — data formats are easy to clone. Stronger as part of the system (#1+#2+#3).',
    },
    {
        'num': 4,
        'name': 'Automated Object Curation with Anti-Paralysis Protocol',
        'priority': 'MEDIUM',
        'value': 'Medium',
        'claim': 'A system for autonomously curating 3D object datasets that enforces decision deadlines with maximum analysis passes and wall-clock timeouts, tiered quality thresholds for auto-approve/reject/flag, persistent hash-based state for interrupt-resume operation, and automatic library registration of approved objects.',
        'novel': 'Anti-paralysis: forced decisions after N passes or T seconds. Quality tiers (auto-approve ≥70%, reject <30%, flag middle). Browsing pace with random intervals. Resume without reprocessing.',
        'prior_art': 'Active learning (Settles 2012) — human-in-loop. AutoML dataset curation — no decision deadlines. ShapeNet tools — no quality gates or anti-paralysis.',
        'notes': 'May be better as trade secret than patent. Method is hard to detect infringement.',
    },
    {
        'num': 5,
        'name': 'Surface Energy Decomposition for Material Interfaces',
        'priority': 'LOW',
        'value': 'Low-Med',
        'claim': 'A method for computing material surface energy by decomposing cohesive energy into electromagnetic and nuclear-mass-dependent components using broken-bond crystal geometry, and applying the result simultaneously to physical simulation and visual rendering of material interfaces.',
        'novel': 'EM vs QCD-scaling decomposition of surface energy. Broken-bond model parameterized by crystal structure and cohesive energy. σ-field sensitivity for extreme environments.',
        'prior_art': 'Broken-bond model (Mackenzie 1962) — well-known in surface science. DFT surface calculations — computational, not real-time. No prior use for rendering.',
        'notes': 'Physics method — harder to patent. The APPLICATION to rendering is the patentable part. Folds into #2.',
    },
]

for i, c in enumerate(candidates):
    row = 5 + i
    ws.row_dimensions[row].height = 90

    ws.cell(row=row, column=1, value=c['num']).font = BODY_FONT
    ws.cell(row=row, column=1).alignment = CENTER

    ws.cell(row=row, column=2, value=c['name']).font = Font(name='Arial', bold=True, size=10)
    ws.cell(row=row, column=2).alignment = WRAP

    priority_cell = ws.cell(row=row, column=3, value=c['priority'])
    priority_cell.font = Font(name='Arial', bold=True, size=10)
    priority_cell.alignment = CENTER
    if c['priority'] == 'HIGH':
        priority_cell.fill = PRIORITY_HIGH
    elif c['priority'] == 'MEDIUM':
        priority_cell.fill = PRIORITY_MED
    else:
        priority_cell.fill = PRIORITY_LOW

    ws.cell(row=row, column=4, value=c['value']).font = BODY_FONT
    ws.cell(row=row, column=4).alignment = CENTER

    ws.cell(row=row, column=5, value=c['claim']).font = BODY_FONT
    ws.cell(row=row, column=5).alignment = WRAP

    ws.cell(row=row, column=6, value=c['novel']).font = BODY_FONT
    ws.cell(row=row, column=6).alignment = WRAP

    ws.cell(row=row, column=7, value=c['prior_art']).font = BODY_FONT
    ws.cell(row=row, column=7).alignment = WRAP

    ws.cell(row=row, column=8, value=c['notes']).font = BLUE_FONT
    ws.cell(row=row, column=8).alignment = WRAP

    for col_idx in range(1, 9):
        ws.cell(row=row, column=col_idx).border = THIN_BORDER

# ── Sheet 2: Filing Strategy ─────────────────────────────────────
ws2 = wb.create_sheet("Filing Strategy")

for col, w in {'A': 5, 'B': 30, 'C': 18, 'D': 18, 'E': 45, 'F': 40}.items():
    ws2.column_dimensions[col].width = w

ws2.merge_cells('A1:F1')
ws2['A1'] = 'Recommended Filing Strategy'
ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='1B2A4A')
ws2.row_dimensions[1].height = 30

ws2.merge_cells('A2:F2')
ws2['A2'] = 'NOT LEGAL ADVICE — Discuss with patent attorney before filing'
ws2['A2'].font = Font(name='Arial', bold=True, italic=True, size=10, color='CC0000')

strategy_headers = ['#', 'Action', 'Timeline', 'Est. Cost', 'Description', 'Dependencies']
for col_idx, h in enumerate(strategy_headers, 1):
    cell = ws2.cell(row=4, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

steps = [
    {
        'num': 1,
        'action': 'File Provisional #1+#2 (Combined System Patent)',
        'timeline': 'Week 1',
        'cost': '$320-$500',
        'desc': 'Provisional patent application covering the push rendering method AND the unified material identity system as a combined system. This is your strongest IP and should be filed BEFORE any publication.',
        'deps': 'None — file immediately',
    },
    {
        'num': 2,
        'action': 'File Provisional #3 (Sigma Signature Format)',
        'timeline': 'Week 1-2',
        'cost': '$320-$500',
        'desc': 'Provisional covering the analytic quadric object representation with physics-aware materials. Can be combined with #1+#2 filing to save costs.',
        'deps': 'Can be combined with Step 1',
    },
    {
        'num': 3,
        'action': 'Consult Patent Attorney',
        'timeline': 'Week 2-4',
        'cost': '$500-$2,000',
        'desc': 'Initial consultation to review provisionals, assess claim strength, and advise on utility patent strategy. Look for IP attorneys with software/physics patent experience.',
        'deps': 'After provisionals filed',
    },
    {
        'num': 4,
        'action': 'Publish Physics Paper (PhysRevD)',
        'timeline': 'Month 2-3',
        'cost': '$0',
        'desc': 'Submit SSBM paper with TOV prediction and chirp mass test. MUST be after provisional filing — publication creates prior art. The paper strengthens the patent by demonstrating the physics foundation.',
        'deps': 'AFTER Step 1 filed',
    },
    {
        'num': 5,
        'action': 'Evaluate #4 (Nagatha) as Trade Secret',
        'timeline': 'Month 2-3',
        'cost': '$0',
        'desc': 'The anti-paralysis curation protocol may be better protected as a trade secret than a patent. Hard to detect infringement. Discuss with attorney.',
        'deps': 'After Step 3 consultation',
    },
    {
        'num': 6,
        'action': 'File Utility Patent (Non-Provisional)',
        'timeline': 'Month 10-12',
        'cost': '$5,000-$15,000',
        'desc': 'Convert strongest provisional(s) to full utility patent within 12-month window. Attorney drafts formal claims based on provisional + any improvements made in the interim.',
        'deps': 'Before 12-month provisional deadline',
    },
]

for i, s in enumerate(steps):
    row = 5 + i
    ws2.row_dimensions[row].height = 60

    ws2.cell(row=row, column=1, value=s['num']).font = BODY_FONT
    ws2.cell(row=row, column=1).alignment = CENTER

    ws2.cell(row=row, column=2, value=s['action']).font = Font(name='Arial', bold=True, size=10)
    ws2.cell(row=row, column=2).alignment = WRAP

    ws2.cell(row=row, column=3, value=s['timeline']).font = BODY_FONT
    ws2.cell(row=row, column=3).alignment = CENTER

    ws2.cell(row=row, column=4, value=s['cost']).font = BLUE_FONT
    ws2.cell(row=row, column=4).alignment = CENTER

    ws2.cell(row=row, column=5, value=s['desc']).font = BODY_FONT
    ws2.cell(row=row, column=5).alignment = WRAP

    ws2.cell(row=row, column=6, value=s['deps']).font = BODY_FONT
    ws2.cell(row=row, column=6).alignment = WRAP

    for col_idx in range(1, 7):
        ws2.cell(row=row, column=col_idx).border = THIN_BORDER

# ── Sheet 3: Technical Evidence ──────────────────────────────────
ws3 = wb.create_sheet("Technical Evidence")

for col, w in {'A': 5, 'B': 28, 'C': 35, 'D': 20, 'E': 40}.items():
    ws3.column_dimensions[col].width = w

ws3.merge_cells('A1:E1')
ws3['A1'] = 'Technical Evidence Inventory'
ws3['A1'].font = Font(name='Arial', bold=True, size=14, color='1B2A4A')
ws3.row_dimensions[1].height = 30

ws3.merge_cells('A2:E2')
ws3['A2'] = 'Code artifacts, test results, and benchmarks supporting patent claims'
ws3['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')

evidence_headers = ['#', 'Artifact', 'File Path', 'Status', 'Supports Claim']
for col_idx, h in enumerate(evidence_headers, 1):
    cell = ws3.cell(row=4, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

evidence = [
    ['1', 'Push renderer (no ray tracing)', 'MatterShaper/mattershaper/render/push.py', '14/14 tests pass', 'Patent #1'],
    ['2', 'Push render benchmark image', 'push_render_coffee_mug.png', 'Visible object, 0 rays', 'Patent #1'],
    ['3', 'Source code audit (no Ray/Hit imports)', 'test_push.py: test_no_ray_objects_used', 'PASS', 'Patent #1'],
    ['4', 'Material identity system', 'MatterShaper/mattershaper/materials/material.py', 'In production', 'Patent #2'],
    ['5', 'Surface energy module', 'local_library/interface/surface.py', '28/28 tests pass', 'Patent #2, #5'],
    ['6', 'Surface test samples (Nagatha feed)', 'MatterShaper/harvest/feed/surface_enriched/', '8/8 materials validated', 'Patent #2'],
    ['7', 'Coffee mug shape signature', 'MatterShaper/object_maps/coffee_mug.shape.json', '12 quadric primitives', 'Patent #3'],
    ['8', 'Coffee mug material map', 'MatterShaper/object_maps/coffee_mug.color.json', '5 physics-aware materials', 'Patent #2, #3'],
    ['9', 'Harvest curator', 'MatterShaper/harvest/harvest_curator.py', 'In production', 'Patent #4'],
    ['10', 'SSBM formula audit', 'local_library/audit.py', '35 formulas, 0 EXTERNAL', 'All patents (foundation)'],
    ['11', 'TOV mass prediction', 'local_library/unsolved.py', 'M_TOV = 2.071 M☉ (within 3σ)', 'Physics paper'],
    ['12', 'Wheeler invariance tests', 'local_library/verify.py', '48/48 exact', 'Physics paper'],
    ['13', 'Scorecard', 'local_library/scorecard.py', '10 SOLVED, 0 DISPROVED', 'Physics paper'],
]

for i, e in enumerate(evidence):
    row = 5 + i
    ws3.row_dimensions[row].height = 22
    for col_idx, val in enumerate(e, 1):
        cell = ws3.cell(row=row, column=col_idx, value=val)
        cell.font = BODY_FONT
        cell.alignment = WRAP if col_idx >= 3 else CENTER
        cell.border = THIN_BORDER
    # Color the status green if passing
    status_cell = ws3.cell(row=row, column=4)
    if 'pass' in status_cell.value.lower() or 'visible' in status_cell.value.lower() or 'exact' in status_cell.value.lower():
        status_cell.fill = PRIORITY_HIGH

# ── Footer notes ─────────────────────────────────────────────────
for sheet in [ws, ws2, ws3]:
    max_row = sheet.max_row + 2
    sheet.merge_cells(f'A{max_row}:H{max_row}' if sheet == ws else f'A{max_row}:F{max_row}')
    footer = sheet.cell(row=max_row, column=1, value='DISCLAIMER: This document is an engineering assessment, not legal advice. Consult a registered patent attorney before filing. Generated by AI assistant.')
    footer.font = Font(name='Arial', italic=True, size=8, color='999999')

out_path = '/sessions/loving-pensive-euler/mnt/quarksum/MatterShaper_Patent_Candidates.xlsx'
wb.save(out_path)
print(f'Saved to {out_path}')
